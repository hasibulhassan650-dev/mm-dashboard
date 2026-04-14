"""
export/exporter.py

Export daily_flow and event DataFrames to CSV or Excel.
Filename pattern: mm-dashboard-YYYYMMDD-{page}.{ext}
Filters applied at export time — never exports full unfiltered dataset.
"""

import io
import logging
from datetime import date

import pandas as pd

logger = logging.getLogger(__name__)

_AMOUNT_COLS = [
    "coupon_inflow_bdt_mill", "principal_inflow_bdt_mill", "total_inflow_bdt_mill",
    "auction_outflow_planned", "auction_outflow_confirmed", "auction_outflow_best",
    "net_borrowing_bdt_mill", "amount_bdt_mill", "principal_bdt_mill",
    "coupon_amount_bdt_mill", "offered_amount_bdt_mill", "accepted_amount_bdt_mill",
    "best_amount_bdt_mill", "outstanding_bdt_mill", "outstanding_used_bdt_mill",
    "cumulative_net_borrowing",
]


def _export_filename(page: str, ext: str) -> str:
    today = date.today().strftime("%Y%m%d")
    return f"mm-dashboard-{today}-{page}.{ext}"


def to_csv_bytes(df: pd.DataFrame, page: str) -> tuple[bytes, str]:
    """Returns (csv_bytes, filename)."""
    buf = io.BytesIO()
    df.to_csv(buf, index=False, encoding="utf-8-sig")
    return buf.getvalue(), _export_filename(page, "csv")


def to_excel_bytes(df: pd.DataFrame, page: str) -> tuple[bytes, str]:
    """
    Returns (excel_bytes, filename).
    Applies number formatting to amount columns.
    """
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Data")
        wb = writer.book
        ws = writer.sheets["Data"]

        # Header formatting
        from openpyxl.styles import PatternFill, Font, Alignment
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Number formatting for amount columns
        amount_fmt = '#,##0.00'
        for col_idx, col_name in enumerate(df.columns, start=1):
            if col_name in _AMOUNT_COLS:
                for row_idx in range(2, ws.max_row + 1):
                    ws.cell(row=row_idx, column=col_idx).number_format = amount_fmt

        # Auto-width
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    return buf.getvalue(), _export_filename(page, "xlsx")


def format_display_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    Format a DataFrame for UI display: comma-format amounts, DD-MMM-YYYY dates.
    Returns a copy — does not mutate input.
    """
    df = df.copy()
    for col in df.columns:
        if col in _AMOUNT_COLS and col in df.columns:
            df[col] = df[col].apply(
                lambda v: f"{v:,.2f}" if pd.notnull(v) else "—"
            )
        if "date" in col.lower() and col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce").dt.strftime("%d-%b-%Y")
    return df
